from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def insertar_componentes(ws, ph_listado, fila_inicio=15):
    fila = fila_inicio
    filas_combinadas = []  # Lista para guardar la fila de inicio de cada PH

    for ph_dato in ph_listado:
        fila_inicio_ph = fila
        filas_combinadas.append(fila_inicio_ph)

        for poligono in ph_dato["poligonos"]:
            for componente in poligono["componentes"]:
                ws[f"B{fila}"] = ph_dato["padron"]
                ws[f"C{fila}"] = ph_dato["unidad"]
                ws[f"D{fila}"] = poligono["poligono"]
                ws[f"E{fila}"] = componente["concepto"]
                ws[f"F{fila}"] = componente["m2"]
                ws[f"G{fila}"] = componente["valor_m2"]
                ws[f"H{fila}"] = componente["coef_antig"]               

                ws[f"I{fila}"] = f"=F{fila}*G{fila}*H{fila}"
                ws[f"M{fila}"] = f"=K{fila}*L{fila}"
                fila += 1

        fila_fin_ph = fila - 1

        combinar_columnas(ws, fila_inicio_ph, fila_fin_ph, columnas=[2, 3])  # B, C
        combinar_columnas(ws, fila_inicio_ph, fila_fin_ph, columnas=list(range(11, 20)))  # K a S

        # Escribir sup_m2 y coef_ajus en la fila de inicio del PH
        ws[f"K{fila_inicio_ph}"] = ph_dato["sup_m2"]
        ws[f"L{fila_inicio_ph}"] = ph_dato["coef_ajus"]

        # Total por PH en columna J
        if fila_fin_ph > fila_inicio_ph:
            ws.merge_cells(start_row=fila_inicio_ph, start_column=10, end_row=fila_fin_ph, end_column=10)
            ws[f"J{fila_inicio_ph}"] = f"=SUM(I{fila_inicio_ph}:I{fila_fin_ph})"
        else:
            ws[f"J{fila_inicio_ph}"] = f"=I{fila_inicio_ph}"

    return filas_combinadas, fila  # fila es la primera fila libre luego de insertar

def combinar_columnas(ws, fila_inicio, fila_fin, columnas):
    for col in columnas:
        if fila_fin > fila_inicio:
            ws.merge_cells(start_row=fila_inicio, start_column=col, end_row=fila_fin, end_column=col)


def escribir_totales_generales(ws, fila_inicio, fila_fin, columnas=range(13, 20)):
    """
    Escribe la suma de cada columna en la fila 'fila_fin', desde fila_inicio hasta fila_fin - 1.
    Por defecto, columnas M (13) a S (19).
    """
    for col in columnas:
        #letra_col = chr(64 + col)
        letra_col = get_column_letter(col)
        ws[f"{letra_col}{fila_fin}"] = f"=SUM({letra_col}{fila_inicio}:{letra_col}{fila_fin - 1})"


def combinar_valores_iguales(ws, col, fila_inicio, fila_fin):
    """
    Combina celdas verticales en la columna 'col' desde 'fila_inicio' hasta 'fila_fin'
    solo cuando los valores son iguales y contiguos.
    """
    letra_col = get_column_letter(col)
    fila_actual = fila_inicio

    while fila_actual <= fila_fin:
        valor_actual = ws[f"{letra_col}{fila_actual}"].value
        fila_grupo_fin = fila_actual

        # Buscar el rango de filas contiguas con el mismo valor
        while (fila_grupo_fin + 1 <= fila_fin and
               ws[f"{letra_col}{fila_grupo_fin + 1}"].value == valor_actual):
            fila_grupo_fin += 1

        # Si hay más de una fila con el mismo valor, combinarlas
        if fila_grupo_fin > fila_actual:
            ws.merge_cells(start_row=fila_actual, start_column=col, end_row=fila_grupo_fin, end_column=col)

        fila_actual = fila_grupo_fin + 1


def calcular_columna_n(ws, valor_tierra, filas_combinadas, fila_final):
    """
    Calcula e inserta en la columna N, para cada PH, la fórmula:
    (valor_tierra / SUMA(M)) * M_fila
    """
    # rango_m = f"M15:M{fila_final - 1}"    
    # for fila in filas_combinadas:
    #     ws[f"N{fila}"] = f"=({valor_tierra}/SUM({rango_m}))*M{fila}"

    sum_M = f"M{fila_final}"    
    for fila in filas_combinadas:
        ws[f"N{fila}"] = f"=({valor_tierra}/{sum_M})*M{fila}"


def calcular_columna_o(ws, filas_combinadas):
    """
    Para cada PH, coloca en la columna O la fórmula:
    =M_fila + N_fila
    """
    for fila in filas_combinadas:
        ws[f"O{fila}"] = f"=J{fila}+N{fila}"

def calcular_columna_p(ws, filas_combinadas, fila_final):
    """
    Calcula e inserta en la columna N, para cada PH, la fórmula:
    (valor_tierra / SUMA(M)) * M_fila
    """
    
    sum_O = f"O{fila_final}"    
    for fila in filas_combinadas:
        ws[f"P{fila}"] = f"=(1000/{sum_O})*O{fila}"

def copiar_columna_n_en_q(ws, filas_combinadas):
    """
    Copia el valor de la columna N en la columna Q para cada PH (una fila por PH).
    """
    for fila in filas_combinadas:
        ws[f"Q{fila}"] = f"=N{fila}"

def calcular_columna_s(ws, filas_combinadas, fila_final):
    """
    Calcula e inserta en la columna N, para cada PH, la fórmula:
    (valor_tierra / SUMA(M)) * M_fila
    """

    sum_P = f"P{fila_final}"    
    for fila in filas_combinadas:
        ws[f"S{fila}"] = f"=(Y17/{sum_P})*P{fila}"

def restar_columna_s_menos_q(ws, filas_combinadas):
    """
    Copia el valor de la columna N en la columna Q para cada PH (una fila por PH).
    """
    for fila in filas_combinadas:
        ws[f"R{fila}"] = f"=S{fila}-Q{fila}"



# =======================
# SCRIPT PRINCIPAL
# =======================

# Cargar planilla base
wb = load_workbook("planilla_base.xlsx")
ws = wb.active

# Datos de ejemplo
ph_listado = [
    {
        "padron": "P-90001",
        "unidad": 5,
        "sup_m2": 50,
        "coef_ajus": 0.9,
        "poligonos": [
            {
                "poligono": "00-01",
                "componentes": [
                    {"concepto": "Cubierta", "m2": 40.0, "valor_m2": 310.0, "coef_antig": 0.85},
                    {"concepto": "Semicub.", "m2": 5.0, "valor_m2": 150.0, "coef_antig": 0.75}
                ]
            },
            {
                "poligono": "01-01",
                "componentes": [
                    {"concepto": "Cubierta", "m2": 30.0, "valor_m2": 320.0, "coef_antig": 0.80},
                    {"concepto": "Cubierta", "m2": 10.0, "valor_m2": 300.0, "coef_antig": 0.90}
                ]
            }
        ]
    },
    {
        "padron": "P-90002",
        "unidad": 6,
        "sup_m2": 120,
        "coef_ajus": 0.8,
        "poligonos": [
            {
                "poligono": "00-02",
                "componentes": [
                    {"concepto": "Cubierta", "m2": 60.0, "valor_m2": 330.0, "coef_antig": 0.80},
                    {"concepto": "Semicub.", "m2": 12.0, "valor_m2": 160.0, "coef_antig": 0.70}
                ]
            }
        ]
    },
    {
        "padron": "P-90003",
        "unidad": 7,
        "sup_m2": 150,
        "coef_ajus": 1,
        "poligonos": [
            {
                "poligono": "00-03",
                "componentes": [
                    {"concepto": "Cubierta", "m2": 25.0, "valor_m2": 290.0, "coef_antig": 0.88},
                    {"concepto": "Galpón", "m2": 15.0, "valor_m2": 140.0, "coef_antig": 0.65}
                ]
            },
            {
                "poligono": "01-03",
                "componentes": [
                    {"concepto": "Pileta", "m2": 20.0, "valor_m2": 220.0, "coef_antig": 0.90}
                ]
            }
        ]
    }
]


# Valores únicos
valor_tierra = 5924.8
valor_mejoras = 209041.58
valor_total = valor_tierra + valor_mejoras
ws["Y15"] = valor_tierra
ws["Y16"] = valor_mejoras
ws["Y17"] = valor_total

# Insertar datos en la planilla
filas_combinadas, fila_fin = insertar_componentes(ws, ph_listado)

# Escribir totales generales
escribir_totales_generales(ws, fila_inicio=15, fila_fin=fila_fin)

# Combinar valores iguales en columna D (Polígono)
combinar_valores_iguales(ws, col=4, fila_inicio=15, fila_fin=fila_fin - 1)

# Calcular columna N
calcular_columna_n(ws, valor_tierra, filas_combinadas, fila_fin)

# Columna O: valor total PH = M + N
calcular_columna_o(ws, filas_combinadas)

# Columna P: calcular porcentual
calcular_columna_p(ws, filas_combinadas, fila_fin)

# Columna Q: igual a columna N
copiar_columna_n_en_q(ws, filas_combinadas)

# Columna S: (valor total / sumatoria de P) * valor P fila
calcular_columna_s(ws, filas_combinadas, fila_fin)

# Columna R: valor total - valor tierra
restar_columna_s_menos_q(ws, filas_combinadas)

# Guardar archivo final
wb.save("planilla_final.xlsx")

# Para control visual en consola (opcional)
#print("Filas combinadas (PH):", filas_combinadas)
#print("Primera fila libre:", fila_fin)

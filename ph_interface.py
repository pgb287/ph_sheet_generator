import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
from typing import List, Dict, Any
from main import insertar_componentes, combinar_columnas, escribir_totales_generales
from main import combinar_valores_iguales, calcular_columna_n, calcular_columna_o
from main import calcular_columna_p, copiar_columna_n_en_q, calcular_columna_s, restar_columna_s_menos_q
from openpyxl import load_workbook


class PHInterface:
    def __init__(self, root):
        self.root = root
        self.root.title("Generador de planillas de PH")
        self.root.geometry("1200x800")
        
        # Datos
        self.ph_listado = []
        # Nuevos campos generales
        self.plano_nro = tk.StringVar()
        self.dpto = tk.StringVar()
        self.localidad = tk.StringVar()
        self.barrio = tk.StringVar()
        self.circ = tk.StringVar()
        self.secc = tk.StringVar()
        self.mza = tk.StringVar()
        self.parc = tk.StringVar()
        self.valor_tierra = tk.DoubleVar()
        self.valor_mejoras = tk.DoubleVar()
        
        self.setup_ui()
        #self.load_sample_data()
    
    def setup_ui(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")
        
        # Configurar grid
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Valores generales
        self.setup_valores_generales(main_frame)
        
        # Frame para PHs
        self.setup_ph_frame(main_frame)
        
        # Botones de acción
        self.setup_action_buttons(main_frame)
    
    def setup_valores_generales(self, parent):
        # Frame para valores generales
        valores_frame = ttk.LabelFrame(parent, text="Valores Generales", padding="5")
        valores_frame.grid(row=0, column=0, columnspan=2, sticky="we", pady=(0, 10))
        
        # Primera fila
        ttk.Label(valores_frame, text="Plano N°:").grid(row=0, column=0, padx=(0, 5), sticky=tk.W)
        ttk.Entry(valores_frame, textvariable=self.plano_nro, width=22).grid(row=0, column=1, padx=(0, 15))
        ttk.Label(valores_frame, text="Dpto:").grid(row=0, column=2, padx=(0, 5), sticky=tk.W)
        ttk.Entry(valores_frame, textvariable=self.dpto, width=22).grid(row=0, column=3, padx=(0, 15))
        ttk.Label(valores_frame, text="Localidad:").grid(row=0, column=4, padx=(0, 5), sticky=tk.W)
        ttk.Entry(valores_frame, textvariable=self.localidad, width=22).grid(row=0, column=5, padx=(0, 15))
        ttk.Label(valores_frame, text="Barrio:").grid(row=0, column=6, padx=(0, 5), sticky=tk.W)
        ttk.Entry(valores_frame, textvariable=self.barrio, width=22).grid(row=0, column=7, padx=(0, 0))
        
        # Segunda fila
        ttk.Label(valores_frame, text="Circ:").grid(row=1, column=0, padx=(0, 5), sticky=tk.W)
        ttk.Entry(valores_frame, textvariable=self.circ, width=22).grid(row=1, column=1, padx=(0, 15))
        ttk.Label(valores_frame, text="Secc:").grid(row=1, column=2, padx=(0, 5), sticky=tk.W)
        ttk.Entry(valores_frame, textvariable=self.secc, width=22).grid(row=1, column=3, padx=(0, 15))
        ttk.Label(valores_frame, text="Mza:").grid(row=1, column=4, padx=(0, 5), sticky=tk.W)
        ttk.Entry(valores_frame, textvariable=self.mza, width=22).grid(row=1, column=5, padx=(0, 15))
        ttk.Label(valores_frame, text="Parc:").grid(row=1, column=6, padx=(0, 5), sticky=tk.W)
        ttk.Entry(valores_frame, textvariable=self.parc, width=22).grid(row=1, column=7, padx=(0, 0))
        
        # Tercera fila
        ttk.Label(valores_frame, text="Valor Tierra:").grid(row=2, column=0, padx=(0, 5), sticky=tk.W)
        ttk.Entry(valores_frame, textvariable=self.valor_tierra, width=22).grid(row=2, column=1, padx=(0, 15))
        ttk.Label(valores_frame, text="Valor Mejoras:").grid(row=2, column=2, padx=(0, 5), sticky=tk.W)
        ttk.Entry(valores_frame, textvariable=self.valor_mejoras, width=22).grid(row=2, column=3, padx=(0, 15))
    
    def setup_ph_frame(self, parent):
        # Frame para PHs
        ph_frame = ttk.LabelFrame(parent, text="Propiedades Horizontales", padding="5")
        ph_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 10))
        ph_frame.columnconfigure(0, weight=1)
        ph_frame.rowconfigure(1, weight=1)
        
        # Botones para PHs
        ph_buttons_frame = ttk.Frame(ph_frame)
        ph_buttons_frame.grid(row=0, column=0, sticky="we", pady=(0, 5))
        
        ttk.Button(ph_buttons_frame, text="Agregar PH", command=self.add_ph).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(ph_buttons_frame, text="Editar PH", command=self.edit_ph).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(ph_buttons_frame, text="Eliminar PH", command=self.delete_ph).pack(side=tk.LEFT)
        
        # Treeview para PHs
        self.ph_tree = ttk.Treeview(ph_frame, columns=("padron", "unidad", "sup_m2", "coef_ajus", "poligonos"), show="tree headings")
        self.ph_tree.grid(row=1, column=0, sticky="nsew")
        
        # Configurar columnas
        self.ph_tree.heading("#0", text="PH")
        self.ph_tree.heading("padron", text="Padrón")
        self.ph_tree.heading("unidad", text="Unidad")
        self.ph_tree.heading("sup_m2", text="Sup. m²")
        self.ph_tree.heading("coef_ajus", text="Coef. Ajuste")
        self.ph_tree.heading("poligonos", text="Polígonos")
        
        self.ph_tree.column("#0", width=100, anchor="center")
        self.ph_tree.column("padron", width=100, anchor="center")
        self.ph_tree.column("unidad", width=80, anchor="center")
        self.ph_tree.column("sup_m2", width=80, anchor="center")
        self.ph_tree.column("coef_ajus", width=100, anchor="center")
        self.ph_tree.column("poligonos", width=100, anchor="center")
        
        # Scrollbar
        ph_scrollbar = ttk.Scrollbar(ph_frame, orient=tk.VERTICAL, command=self.ph_tree.yview)
        ph_scrollbar.grid(row=1, column=1, sticky="ns")
        self.ph_tree.configure(yscrollcommand=ph_scrollbar.set)
        
        # Frame para componentes
        self.setup_componentes_frame(parent)
    
    def setup_componentes_frame(self, parent):
        # Frame para componentes
        comp_frame = ttk.LabelFrame(parent, text="Componentes del PH Seleccionado", padding="5")
        comp_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", pady=(0, 10))
        comp_frame.columnconfigure(0, weight=1)
        comp_frame.rowconfigure(1, weight=1)
        
        # Botones para componentes
        comp_buttons_frame = ttk.Frame(comp_frame)
        comp_buttons_frame.grid(row=0, column=0, sticky="we", pady=(0, 5))
        
        ttk.Button(comp_buttons_frame, text="Agregar Polígono", command=self.add_poligono).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(comp_buttons_frame, text="Agregar Componente", command=self.add_componente).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(comp_buttons_frame, text="Eliminar Componente", command=self.delete_componente).pack(side=tk.LEFT)
        
        # Treeview para componentes
        self.comp_tree = ttk.Treeview(comp_frame, columns=("poligono", "concepto", "m2", "valor_m2", "coef_antig"), show="tree headings")
        self.comp_tree.grid(row=1, column=0, sticky="nsew")
        
        # Configurar columnas
        self.comp_tree.heading("#0", text="Componente")
        self.comp_tree.heading("poligono", text="Polígono")
        self.comp_tree.heading("concepto", text="Concepto")
        self.comp_tree.heading("m2", text="m²")
        self.comp_tree.heading("valor_m2", text="Valor/m²")
        self.comp_tree.heading("coef_antig", text="Coef. Antig.")
        
        self.comp_tree.column("#0", width=150, anchor="center")
        self.comp_tree.column("poligono", width=100, anchor="center")
        self.comp_tree.column("concepto", width=120, anchor="center")
        self.comp_tree.column("m2", width=80, anchor="center")
        self.comp_tree.column("valor_m2", width=100, anchor="center")
        self.comp_tree.column("coef_antig", width=100, anchor="center")
        
        # Scrollbar
        comp_scrollbar = ttk.Scrollbar(comp_frame, orient=tk.VERTICAL, command=self.comp_tree.yview)
        comp_scrollbar.grid(row=1, column=1, sticky="ns")
        self.comp_tree.configure(yscrollcommand=comp_scrollbar.set)
        
        # Eventos
        self.ph_tree.bind('<<TreeviewSelect>>', self.on_ph_select)
    
    def setup_action_buttons(self, parent):
        # Frame para botones de acción
        action_frame = ttk.Frame(parent)
        action_frame.grid(row=3, column=0, columnspan=2, pady=(10, 0))
        
        # Eliminar botones de Guardar y Cargar datos
        # ttk.Button(action_frame, text="Guardar Datos", command=self.save_data).pack(side=tk.LEFT, padx=(0, 5))
        # ttk.Button(action_frame, text="Cargar Datos", command=self.load_data).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(action_frame, text="Generar Excel", command=self.generate_excel).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(action_frame, text="Limpiar Todo", command=self.clear_all).pack(side=tk.LEFT)
    
    def load_sample_data(self):
        """Cargar datos de ejemplo"""
        self.ph_listado = [
            {
                "padron": "P-90001",
                "unidad": 5,
                "sup_m2": 50,
                "coef_ajus": 0.9,
                "poligonos": [
                    {
                        "poligono": "00-01",
                        "componentes": [
                            {"concepto": "Cubierta", "m2": 40.0, "valor_m2": 310.0, "coef_antig": 0.85},
                            {"concepto": "Semicub.", "m2": 5.0, "valor_m2": 150.0, "coef_antig": 0.75}
                        ]
                    }
                ]
            }
        ]
        self.refresh_ph_tree()
    
    def refresh_ph_tree(self):
        """Actualizar el árbol de PHs"""
        for item in self.ph_tree.get_children():
            self.ph_tree.delete(item)
        
        for i, ph in enumerate(self.ph_listado):
            poligonos_count = len(ph["poligonos"])
            self.ph_tree.insert("", "end", text=f"PH {i+1}", values=(
                ph["padron"], ph["unidad"], ph["sup_m2"], ph["coef_ajus"], poligonos_count
            ))
    
    def refresh_componentes_tree(self, ph_index):
        """Actualizar el árbol de componentes para un PH específico"""
        for item in self.comp_tree.get_children():
            self.comp_tree.delete(item)
        
        if ph_index >= 0 and ph_index < len(self.ph_listado):
            ph = self.ph_listado[ph_index]
            for poligono in ph["poligonos"]:
                poligono_item = self.comp_tree.insert("", "end", text=poligono["poligono"], values=(
                    poligono["poligono"], "", "", "", ""
                ))
                
                for componente in poligono["componentes"]:
                    self.comp_tree.insert(poligono_item, "end", text=componente["concepto"], values=(
                        poligono["poligono"], componente["concepto"], componente["m2"], 
                        componente["valor_m2"], componente["coef_antig"]
                    ))
    
    def on_ph_select(self, event):
        """Manejar selección de PH"""
        selection = self.ph_tree.selection()
        if selection:
            item = self.ph_tree.item(selection[0])
            ph_index = self.ph_tree.index(selection[0])
            self.refresh_componentes_tree(ph_index)
    
    def add_ph(self):
        """Agregar nuevo PH"""
        dialog = PHDialog(self.root, title="Agregar PH")
        if dialog.result:
            self.ph_listado.append(dialog.result)
            self.refresh_ph_tree()
    
    def edit_ph(self):
        """Editar PH seleccionado"""
        selection = self.ph_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Selecciona un PH para editar")
            return
        
        ph_index = self.ph_tree.index(selection[0])
        dialog = PHDialog(self.root, title="Editar PH", ph_data=self.ph_listado[ph_index])
        if dialog.result:
            self.ph_listado[ph_index] = dialog.result
            self.refresh_ph_tree()
    
    def delete_ph(self):
        """Eliminar PH seleccionado"""
        selection = self.ph_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Selecciona un PH para eliminar")
            return
        
        if messagebox.askyesno("Confirmar", "¿Estás seguro de que quieres eliminar este PH?"):
            ph_index = self.ph_tree.index(selection[0])
            del self.ph_listado[ph_index]
            self.refresh_ph_tree()
            self.refresh_componentes_tree(-1)
    
    def add_poligono(self):
        """Agregar polígono al PH seleccionado"""
        selection = self.ph_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Selecciona un PH primero")
            return
        
        ph_index = self.ph_tree.index(selection[0])
        dialog = PoligonoDialog(self.root, title="Agregar Polígono")
        if dialog.result:
            self.ph_listado[ph_index]["poligonos"].append(dialog.result)
            self.refresh_componentes_tree(ph_index)
    
    def add_componente(self):
        """Agregar componente al polígono seleccionado"""
        poligono_selection = self.comp_tree.selection()
        if not poligono_selection:
            messagebox.showwarning("Advertencia", "Selecciona un polígono primero")
            return
        
        # Encontrar el PH y polígono correspondiente
        ph_selection = self.ph_tree.selection()
        if not ph_selection:
            return
        
        ph_index = self.ph_tree.index(ph_selection[0])
        poligono_item = self.comp_tree.item(poligono_selection[0])
        
        if poligono_item["text"] in ["Cubierta", "Semicub.", "Galpón", "Pileta"]:
            # Es un componente, necesitamos el polígono padre
            parent = self.comp_tree.parent(poligono_selection[0])
            if parent:
                poligono_nombre = self.comp_tree.item(parent)["text"]
            else:
                return
        else:
            poligono_nombre = poligono_item["text"]
        
        dialog = ComponenteDialog(self.root, title="Agregar Componente")
        if dialog.result:
            # Encontrar el polígono correcto
            for poligono in self.ph_listado[ph_index]["poligonos"]:
                if poligono["poligono"] == poligono_nombre:
                    poligono["componentes"].append(dialog.result)
                    break
            
            self.refresh_componentes_tree(ph_index)
    
    def delete_componente(self):
        """Eliminar componente seleccionado"""
        selection = self.comp_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Selecciona un componente para eliminar")
            return
        
        item = self.comp_tree.item(selection[0])
        if item["text"] in ["Cubierta", "Semicub.", "Galpón", "Pileta"]:
            # Es un componente
            if messagebox.askyesno("Confirmar", "¿Eliminar este componente?"):
                # Encontrar y eliminar el componente
                ph_selection = self.ph_tree.selection()
                if ph_selection:
                    ph_index = self.ph_tree.index(ph_selection[0])
                    parent = self.comp_tree.parent(selection[0])
                    if parent:
                        poligono_nombre = self.comp_tree.item(parent)["text"]
                        
                        for poligono in self.ph_listado[ph_index]["poligonos"]:
                            if poligono["poligono"] == poligono_nombre:
                                # Encontrar el componente por concepto
                                for i, componente in enumerate(poligono["componentes"]):
                                    if componente["concepto"] == item["text"]:
                                        del poligono["componentes"][i]
                                        break
                                break
                        
                        self.refresh_componentes_tree(ph_index)
    
    def generate_excel(self):
        """Generar archivo Excel"""
        if not self.ph_listado:
            messagebox.showwarning("Advertencia", "No hay datos para generar el Excel")
            return
        
        try:
            # Cargar planilla base
            wb = load_workbook("planilla_base.xlsx", data_only=True)
            ws = wb.active
            if ws is None:
                ws = wb.worksheets[0] if wb.worksheets else None
                if ws is None:
                    raise ValueError("No se encontró ninguna hoja de trabajo")
            
            # Asignar los nuevos valores a las celdas correspondientes
            ws["O2"] = f"NÚMERO: {self.plano_nro.get()}"
            ws["B7"] = f"DEPARTAMENTO: {self.dpto.get()}"                       
            ws["H7"] = f"LOCALIDAD: {self.localidad.get()}"
            ws["L7"] = f"BARRIOS: {self.barrio.get()}"
            ws["P7"] = f"CIRCUNS: {self.circ.get()}"
            ws["R7"] = f"SECC: {self.secc.get()}"
            ws["B10"] = f"MANZANA: {self.mza.get()}"
            ws["B11"] = f"PARCELA: {self.parc.get()}"
                      
                       
            # Valores únicos
            valor_tierra = self.valor_tierra.get()
            valor_mejoras = self.valor_mejoras.get()
            valor_total = valor_tierra + valor_mejoras
            ws["Y15"] = valor_tierra
            ws["Y16"] = valor_mejoras
            ws["Y17"] = valor_total
            
            # Insertar datos en la planilla
            filas_combinadas, fila_fin = insertar_componentes(ws, self.ph_listado)
            
            # Escribir totales generales
            escribir_totales_generales(ws, fila_inicio=15, fila_fin=fila_fin)
            
            # Combinar valores iguales en columna D (Polígono)
            combinar_valores_iguales(ws, col=4, fila_inicio=15, fila_fin=fila_fin - 1)
            
            # Calcular columnas
            calcular_columna_n(ws, valor_tierra, filas_combinadas, fila_fin)
            calcular_columna_o(ws, filas_combinadas)
            calcular_columna_p(ws, filas_combinadas, fila_fin)
            copiar_columna_n_en_q(ws, filas_combinadas)
            calcular_columna_s(ws, filas_combinadas, fila_fin)
            restar_columna_s_menos_q(ws, filas_combinadas)
            
            # Guardar archivo final
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if filename:
                wb.save(filename)
                messagebox.showinfo("Éxito", "Archivo Excel generado correctamente")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar Excel: {e}")
    
    def clear_all(self):
        """Limpiar todos los datos"""
        if messagebox.askyesno("Confirmar", "¿Estás seguro de que quieres limpiar todos los datos?"):
            self.ph_listado = []
            self.valor_tierra.set(5924.8)
            self.valor_mejoras.set(209041.58)
            self.refresh_ph_tree()
            self.refresh_componentes_tree(-1)


class PHDialog:
    def __init__(self, parent, title, ph_data=None):
        self.result = None
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("400x300")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Variables
        self.padron = tk.StringVar(value=ph_data["padron"] if ph_data else "")
        self.unidad = tk.IntVar(value=ph_data["unidad"] if ph_data else 1)
        self.sup_m2 = tk.DoubleVar(value=ph_data["sup_m2"] if ph_data else 0.0)
        self.coef_ajus = tk.DoubleVar(value=ph_data["coef_ajus"] if ph_data else 1.0)
        
        self.setup_ui()
        self.dialog.wait_window()
    
    def setup_ui(self):
        # Frame principal
        main_frame = ttk.Frame(self.dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Campos
        ttk.Label(main_frame, text="Padrón:").grid(row=0, column=0, sticky=tk.W, pady=2)
        ttk.Entry(main_frame, textvariable=self.padron, width=30).grid(row=0, column=1, sticky="we", padx=(5, 0), pady=2)
        
        ttk.Label(main_frame, text="Unidad:").grid(row=1, column=0, sticky=tk.W, pady=2)
        ttk.Entry(main_frame, textvariable=self.unidad, width=30).grid(row=1, column=1, sticky="we", padx=(5, 0), pady=2)
        
        ttk.Label(main_frame, text="Sup. m²:").grid(row=2, column=0, sticky=tk.W, pady=2)
        ttk.Entry(main_frame, textvariable=self.sup_m2, width=30).grid(row=2, column=1, sticky="we", padx=(5, 0), pady=2)
        
        ttk.Label(main_frame, text="Coef. Ajuste:").grid(row=3, column=0, sticky=tk.W, pady=2)
        ttk.Entry(main_frame, textvariable=self.coef_ajus, width=30).grid(row=3, column=1, sticky="we", padx=(5, 0), pady=2)
        
        # Botones
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=(20, 0))
        
        ttk.Button(button_frame, text="Aceptar", command=self.accept).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="Cancelar", command=self.cancel).pack(side=tk.LEFT)
        
        # Configurar grid
        main_frame.columnconfigure(1, weight=1)
    
    def accept(self):
        """Aceptar y crear resultado"""
        self.result = {
            "padron": self.padron.get(),
            "unidad": self.unidad.get(),
            "sup_m2": self.sup_m2.get(),
            "coef_ajus": self.coef_ajus.get(),
            "poligonos": []
        }
        self.dialog.destroy()
    
    def cancel(self):
        """Cancelar"""
        self.dialog.destroy()


class PoligonoDialog:
    def __init__(self, parent, title):
        self.result = None
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("300x150")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Variables
        self.poligono = tk.StringVar(value="00-01")
        
        self.setup_ui()
        self.dialog.wait_window()
    
    def setup_ui(self):
        # Frame principal
        main_frame = ttk.Frame(self.dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Campos
        ttk.Label(main_frame, text="Polígono:").grid(row=0, column=0, sticky=tk.W, pady=2)
        ttk.Entry(main_frame, textvariable=self.poligono, width=20).grid(row=0, column=1, sticky="we", padx=(5, 0), pady=2)
        
        # Botones
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=1, column=0, columnspan=2, pady=(20, 0))
        
        ttk.Button(button_frame, text="Aceptar", command=self.accept).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="Cancelar", command=self.cancel).pack(side=tk.LEFT)
        
        # Configurar grid
        main_frame.columnconfigure(1, weight=1)
    
    def accept(self):
        """Aceptar y crear resultado"""
        self.result = {
            "poligono": self.poligono.get(),
            "componentes": []
        }
        self.dialog.destroy()
    
    def cancel(self):
        """Cancelar"""
        self.dialog.destroy()


class ComponenteDialog:
    def __init__(self, parent, title):
        self.result = None
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("400x250")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Variables
        self.concepto = tk.StringVar(value="Cubierta")
        self.m2 = tk.DoubleVar(value=0.0)
        self.valor_m2 = tk.DoubleVar(value=0.0)
        self.coef_antig = tk.DoubleVar(value=1.0)
        
        self.setup_ui()
        self.dialog.wait_window()
    
    def setup_ui(self):
        # Frame principal
        main_frame = ttk.Frame(self.dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Campos
        ttk.Label(main_frame, text="Concepto:").grid(row=0, column=0, sticky=tk.W, pady=2)
        concepto_combo = ttk.Combobox(main_frame, textvariable=self.concepto, values=["Cubierta", "Semicub.", "Galpón", "Pileta"], width=27)
        concepto_combo.grid(row=0, column=1, sticky="we", padx=(5, 0), pady=2)
        
        ttk.Label(main_frame, text="m²:").grid(row=1, column=0, sticky=tk.W, pady=2)
        ttk.Entry(main_frame, textvariable=self.m2, width=30).grid(row=1, column=1, sticky="we", padx=(5, 0), pady=2)
        
        ttk.Label(main_frame, text="Valor/m²:").grid(row=2, column=0, sticky=tk.W, pady=2)
        ttk.Entry(main_frame, textvariable=self.valor_m2, width=30).grid(row=2, column=1, sticky="we", padx=(5, 0), pady=2)
        
        ttk.Label(main_frame, text="Coef. Antigüedad:").grid(row=3, column=0, sticky=tk.W, pady=2)
        ttk.Entry(main_frame, textvariable=self.coef_antig, width=30).grid(row=3, column=1, sticky="we", padx=(5, 0), pady=2)
        
        # Botones
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=(20, 0))
        
        ttk.Button(button_frame, text="Aceptar", command=self.accept).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="Cancelar", command=self.cancel).pack(side=tk.LEFT)
        
        # Configurar grid
        main_frame.columnconfigure(1, weight=1)
    
    def accept(self):
        """Aceptar y crear resultado"""
        self.result = {
            "concepto": self.concepto.get(),
            "m2": self.m2.get(),
            "valor_m2": self.valor_m2.get(),
            "coef_antig": self.coef_antig.get()
        }
        self.dialog.destroy()
    
    def cancel(self):
        """Cancelar"""
        self.dialog.destroy()


def main():
    root = tk.Tk()
    app = PHInterface(root)
    root.mainloop()


if __name__ == "__main__":
    main() 
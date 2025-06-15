from openpyxl import load_workbook

def insertar_componentes(ws, ph_listado, fila_inicio=15):
    fila = fila_inicio
    for ph_dato in ph_listado:
        fila_inicio_ph = fila

        for poligono in ph_dato["poligonos"]:
            for componente in poligono["componentes"]:
                ws[f"B{fila}"] = ph_dato["padron"]
                ws[f"C{fila}"] = ph_dato["unidad"]
                ws[f"D{fila}"] = poligono["poligono"]
                ws[f"E{fila}"] = componente["concepto"]
                ws[f"F{fila}"] = componente["m2"]
                ws[f"G{fila}"] = componente["valor_m2"]
                ws[f"H{fila}"] = componente["coef_antig"]
                # Valores nuevos sup_m2 y coef_ajus (ejemplo)
                ws[f"K{fila}"] = componente.get("sup_m2", 1)    # Si no existe, 1 por defecto
                ws[f"L{fila}"] = componente.get("coef_ajus", 1) # Si no existe, 1 por defecto

                ws[f"I{fila}"] = f"=F{fila}*G{fila}*H{fila}"  # cálculo original
                ws[f"M{fila}"] = f"=K{fila}*L{fila}"          # nuevo cálculo solicitado
                fila += 1

        fila_fin_ph = fila - 1

        # Combinar columnas según la lógica previa (se mantiene igual)
        combinar_columnas(ws, fila_inicio_ph, fila_fin_ph, columnas=[2,3,4]) # B, C, D
        combinar_columnas(ws, fila_inicio_ph, fila_fin_ph, columnas=list(range(11, 20))) # K a S (columnas 11 a 19)

        # Insertar total en columna J (col 10)
        if fila_fin_ph > fila_inicio_ph:
            ws.merge_cells(start_row=fila_inicio_ph, start_column=10, end_row=fila_fin_ph, end_column=10)
            ws[f"J{fila_inicio_ph}"] = f"=SUM(I{fila_inicio_ph}:I{fila_fin_ph})"
        else:
            ws[f"J{fila_inicio_ph}"] = f"=I{fila_inicio_ph}"

def combinar_columnas(ws, fila_inicio, fila_fin, columnas):
    for col in columnas:
        if fila_fin > fila_inicio:
            ws.merge_cells(start_row=fila_inicio, start_column=col, end_row=fila_fin, end_column=col)

# Datos ejemplo con sup_m2 y coef_ajus para cada componente
ph_listado = [
    {
        "padron": "P-90001",
        "unidad": 5,
        "poligonos": [
            {
                "poligono": "00-01",
                "componentes": [
                    {"concepto": "Cubierta", "m2": 40.0, "valor_m2": 310.0, "coef_antig": 0.85, "sup_m2": 50, "coef_ajus": 0.9},
                    {"concepto": "Semicub.", "m2": 5.0, "valor_m2": 150.0, "coef_antig": 0.75, "sup_m2": 7, "coef_ajus": 0.7}
                ]
            },
            {
                "poligono": "00-02",
                "componentes": [
                    {"concepto": "Cubierta", "m2": 30.0, "valor_m2": 320.0, "coef_antig": 0.80, "sup_m2": 35, "coef_ajus": 0.85},
                    {"concepto": "Cubierta", "m2": 10.0, "valor_m2": 300.0, "coef_antig": 0.90, "sup_m2": 12, "coef_ajus": 0.95}
                ]
            }
        ]
    },
    {
        "padron": "P-90002",
        "unidad": 6,
        "poligonos": [
            {
                "poligono": "00-01",
                "componentes": [
                    {"concepto": "Cubierta", "m2": 60.0, "valor_m2": 330.0, "coef_antig": 0.80, "sup_m2": 65, "coef_ajus": 0.92},
                    {"concepto": "Semicub.", "m2": 12.0, "valor_m2": 160.0, "coef_antig": 0.70, "sup_m2": 14, "coef_ajus": 0.75}
                ]
            }
        ]
    },
    {
        "padron": "P-90003",
        "unidad": 7,
        "poligonos": [
            {
                "poligono": "00-01",
                "componentes": [
                    {"concepto": "Cubierta", "m2": 25.0, "valor_m2": 290.0, "coef_antig": 0.88, "sup_m2": 28, "coef_ajus": 0.87},
                    {"concepto": "Galpón", "m2": 15.0, "valor_m2": 140.0, "coef_antig": 0.65, "sup_m2": 17, "coef_ajus": 0.62}
                ]
            },
            {
                "poligono": "00-02",
                "componentes": [
                    {"concepto": "Pileta", "m2": 20.0, "valor_m2": 220.0, "coef_antig": 0.90, "sup_m2": 22, "coef_ajus": 0.89}
                ]
            }
        ]
    }
]

# Cargar planilla base
wb = load_workbook("ph_prueba.xlsx")
ws = wb.active

# Insertar datos
insertar_componentes(ws, ph_listado)

# Guardar archivo
wb.save("ph_varios_con_totales_y_sup_coef.xlsx")
